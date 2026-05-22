import io
from datetime import datetime
from database.connection import DatabaseConnection
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF


class ReportsService:

    @staticmethod
    def _fetch_report_data(user_id, role_name):
        db = DatabaseConnection()
        try:
            tasks = db.execute_query("""
                SELECT t.*, u.full_name as assigned_to_name, p.project_name
                FROM Tasks t
                LEFT JOIN Users u ON t.assigned_to = u.user_id
                LEFT JOIN Projects p ON t.project_id = p.project_id
                ORDER BY t.created_at DESC
            """)
            expenses = db.execute_query("""
                SELECT e.*, u.full_name as employee_name, ap.full_name as approved_by_name
                FROM Expenses e
                LEFT JOIN Users u ON e.employee_id = u.user_id
                LEFT JOIN Users ap ON e.approved_by = ap.user_id
                ORDER BY e.submitted_date DESC
            """)
            projects = db.execute_query("""
                SELECT p.*, u.full_name as manager_name
                FROM Projects p
                LEFT JOIN Users u ON p.manager_id = u.user_id
                ORDER BY p.created_at DESC
            """)
            return {'success': True, 'tasks': tasks, 'expenses': expenses, 'projects': projects}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    @staticmethod
    def export_excel(user_id, role_name):
        data = ReportsService._fetch_report_data(user_id, role_name)
        if not data['success']:
            return None, data['message']

        tasks = data['tasks']
        expenses = data['expenses']
        projects = data['projects']

        wb = Workbook()

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4E73DF', end_color='4E73DF', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def style_header(ws, row, cols):
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

        def auto_width(ws, cols):
            for col in range(1, cols + 1):
                max_len = 0
                letter = get_column_letter(col)
                for cell in ws[letter]:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[letter].width = min(max_len + 3, 40)

        # --- Sheet 1: Summary ---
        ws = wb.active
        ws.title = 'Summary'
        task_status_counts = {'Pending': 0, 'In Progress': 0, 'Completed': 0, 'Blocked': 0}
        for t in tasks:
            s = t.get('status', 'Pending')
            if s in task_status_counts:
                task_status_counts[s] += 1
        expense_status_counts = {'Pending': 0, 'Approved': 0, 'Rejected': 0}
        for e in expenses:
            s = e.get('approval_status', 'Pending')
            if s in expense_status_counts:
                expense_status_counts[s] += 1
        project_status_counts = {'Planned': 0, 'Active': 0, 'Completed': 0, 'On Hold': 0}
        for p in projects:
            s = p.get('status', 'Planned')
            if s in project_status_counts:
                project_status_counts[s] += 1
        category_totals = {}
        for e in expenses:
            cat = e.get('category', 'Other')
            category_totals[cat] = category_totals.get(cat, 0) + float(e.get('amount', 0))

        ws['A1'] = 'TaskFlow - Management System Report'
        ws['A1'].font = Font(bold=True, size=14, color='4E73DF')
        ws.merge_cells('A1:B1')
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws['A2'].font = Font(italic=True, size=10, color='666666')

        summary_data = [
            ('Total Tasks', len(tasks)),
            ('Completed Tasks', task_status_counts['Completed']),
            ('Pending Tasks', task_status_counts['Pending']),
            ('In Progress Tasks', task_status_counts['In Progress']),
            ('Total Expenses', len(expenses)),
            ('Approved Expenses', expense_status_counts['Approved']),
            ('Pending Expenses', expense_status_counts['Pending']),
            ('Rejected Expenses', expense_status_counts['Rejected']),
            ('Total Projects', len(projects)),
            ('Active Projects', project_status_counts['Active']),
            ('Completed Projects', project_status_counts['Completed']),
        ]
        for i, (label, value) in enumerate(summary_data, start=4):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)
            ws.cell(row=i, column=1).border = thin_border
            ws.cell(row=i, column=2).border = thin_border
        style_header(ws, 3, 2)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

        # --- Sheet 2: Tasks ---
        ws2 = wb.create_sheet('Tasks')
        headers = ['ID', 'Title', 'Description', 'Status', 'Priority', 'Deadline', 'Assigned To', 'Project', 'Created']
        for col, h in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=h)
        style_header(ws2, 1, len(headers))
        for i, t in enumerate(tasks, start=2):
            ws2.cell(row=i, column=1, value=t.get('task_id'))
            ws2.cell(row=i, column=2, value=t.get('title'))
            ws2.cell(row=i, column=3, value=(t.get('description') or '')[:100])
            ws2.cell(row=i, column=4, value=t.get('status'))
            ws2.cell(row=i, column=5, value=t.get('priority'))
            ws2.cell(row=i, column=6, value=str(t.get('deadline') or ''))
            ws2.cell(row=i, column=7, value=t.get('assigned_to_name') or '')
            ws2.cell(row=i, column=8, value=t.get('project_name') or '')
            ws2.cell(row=i, column=9, value=str(t.get('created_at') or ''))
            for col in range(1, len(headers) + 1):
                ws2.cell(row=i, column=col).border = thin_border
        auto_width(ws2, len(headers))

        # --- Sheet 3: Expenses ---
        ws3 = wb.create_sheet('Expenses')
        headers = ['ID', 'Description', 'Category', 'Amount', 'Status', 'Submitted By', 'Approved By', 'Date']
        for col, h in enumerate(headers, 1):
            ws3.cell(row=1, column=col, value=h)
        style_header(ws3, 1, len(headers))
        for i, e in enumerate(expenses, start=2):
            ws3.cell(row=i, column=1, value=e.get('expense_id'))
            ws3.cell(row=i, column=2, value=e.get('description') or '')
            ws3.cell(row=i, column=3, value=e.get('category'))
            ws3.cell(row=i, column=4, value=float(e.get('amount', 0)))
            ws3.cell(row=i, column=5, value=e.get('approval_status'))
            ws3.cell(row=i, column=6, value=e.get('employee_name') or '')
            ws3.cell(row=i, column=7, value=e.get('approved_by_name') or '')
            ws3.cell(row=i, column=8, value=str(e.get('submitted_date') or ''))
            for col in range(1, len(headers) + 1):
                ws3.cell(row=i, column=col).border = thin_border
        auto_width(ws3, len(headers))

        # --- Sheet 4: Projects ---
        ws4 = wb.create_sheet('Projects')
        headers = ['ID', 'Name', 'Description', 'Status', 'Manager', 'Start Date', 'End Date', 'Created']
        for col, h in enumerate(headers, 1):
            ws4.cell(row=1, column=col, value=h)
        style_header(ws4, 1, len(headers))
        for i, p in enumerate(projects, start=2):
            ws4.cell(row=i, column=1, value=p.get('project_id'))
            ws4.cell(row=i, column=2, value=p.get('project_name'))
            ws4.cell(row=i, column=3, value=(p.get('description') or '')[:100])
            ws4.cell(row=i, column=4, value=p.get('status'))
            ws4.cell(row=i, column=5, value=p.get('manager_name') or '')
            ws4.cell(row=i, column=6, value=str(p.get('start_date') or ''))
            ws4.cell(row=i, column=7, value=str(p.get('end_date') or ''))
            ws4.cell(row=i, column=8, value=str(p.get('created_at') or ''))
            for col in range(1, len(headers) + 1):
                ws4.cell(row=i, column=col).border = thin_border
        auto_width(ws4, len(headers))

        # --- Sheet 5: Category Breakdown ---
        ws5 = wb.create_sheet('Category Breakdown')
        headers = ['Category', 'Total Amount']
        for col, h in enumerate(headers, 1):
            ws5.cell(row=1, column=col, value=h)
        style_header(ws5, 1, len(headers))
        for i, (cat, amt) in enumerate(sorted(category_totals.items()), start=2):
            ws5.cell(row=i, column=1, value=cat)
            ws5.cell(row=i, column=2, value=round(amt, 2))
            ws5.cell(row=i, column=1).border = thin_border
            ws5.cell(row=i, column=2).border = thin_border
        ws5.column_dimensions['A'].width = 25
        ws5.column_dimensions['B'].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, None

    @staticmethod
    def export_pdf(user_id, role_name):
        data = ReportsService._fetch_report_data(user_id, role_name)
        if not data['success']:
            return None, data['message']

        tasks = data['tasks']
        expenses = data['expenses']
        projects = data['projects']

        task_status_counts = {'Pending': 0, 'In Progress': 0, 'Completed': 0, 'Blocked': 0}
        for t in tasks:
            s = t.get('status', 'Pending')
            if s in task_status_counts:
                task_status_counts[s] += 1
        expense_status_counts = {'Pending': 0, 'Approved': 0, 'Rejected': 0}
        for e in expenses:
            s = e.get('approval_status', 'Pending')
            if s in expense_status_counts:
                expense_status_counts[s] += 1
        project_status_counts = {'Planned': 0, 'Active': 0, 'Completed': 0, 'On Hold': 0}
        for p in projects:
            s = p.get('status', 'Planned')
            if s in project_status_counts:
                project_status_counts[s] += 1
        category_totals = {}
        for e in expenses:
            cat = e.get('category', 'Other')
            category_totals[cat] = category_totals.get(cat, 0) + float(e.get('amount', 0))

        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()

        pdf.set_font('Helvetica', 'B', 20)
        pdf.set_text_color(78, 115, 223)
        pdf.cell(0, 15, 'TaskFlow - Management System Report', new_x='LMARGIN', new_y='NEXT', align='C')
        pdf.set_font('Helvetica', '', 10)
        pdf.set_text_color(102, 102, 102)
        pdf.cell(0, 8, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', new_x='LMARGIN', new_y='NEXT', align='C')
        pdf.ln(5)

        def add_table_section(title, headers, rows, col_widths):
            pdf.set_font('Helvetica', 'B', 13)
            pdf.set_text_color(78, 115, 223)
            pdf.cell(0, 10, title, new_x='LMARGIN', new_y='NEXT')
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_fill_color(78, 115, 223)
            pdf.set_text_color(255, 255, 255)
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 7, h, border=1, fill=True, align='C')
            pdf.ln()
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(0, 0, 0)
            fill = False
            for row in rows:
                if pdf.get_y() > 260:
                    pdf.add_page()
                    pdf.set_font('Helvetica', 'B', 8)
                    pdf.set_fill_color(78, 115, 223)
                    pdf.set_text_color(255, 255, 255)
                    for i, h in enumerate(headers):
                        pdf.cell(col_widths[i], 7, h, border=1, fill=True, align='C')
                    pdf.ln()
                    pdf.set_font('Helvetica', '', 8)
                    pdf.set_text_color(0, 0, 0)
                for i, val in enumerate(row):
                    pdf.cell(col_widths[i], 6, str(val)[:40], border=1, fill=fill, align='C')
                pdf.ln()
                fill = not fill
            pdf.ln(5)

        # Summary table
        sum_headers = ['Metric', 'Value']
        sum_rows = [
            ('Total Tasks', str(len(tasks))),
            ('Completed', str(task_status_counts['Completed'])),
            ('Pending', str(task_status_counts['Pending'])),
            ('In Progress', str(task_status_counts['In Progress'])),
            ('Total Expenses', str(len(expenses))),
            ('Approved', str(expense_status_counts['Approved'])),
            ('Pending Expenses', str(expense_status_counts['Pending'])),
            ('Rejected', str(expense_status_counts['Rejected'])),
            ('Total Projects', str(len(projects))),
            ('Active', str(project_status_counts['Active'])),
            ('Completed Projects', str(project_status_counts['Completed'])),
        ]
        add_table_section('Summary Report', sum_headers, sum_rows, [60, 40])

        # Task status breakdown
        t_headers = ['Status', 'Count']
        t_rows = [[k, str(v)] for k, v in task_status_counts.items()]
        add_table_section('Task Status Distribution', t_headers, t_rows, [60, 40])

        # Expense status breakdown
        e_headers = ['Status', 'Count']
        e_rows = [[k, str(v)] for k, v in expense_status_counts.items()]
        add_table_section('Expense Status Overview', e_headers, e_rows, [60, 40])

        # Project status breakdown
        p_headers = ['Status', 'Count']
        p_rows = [[k, str(v)] for k, v in project_status_counts.items()]
        add_table_section('Project Status Overview', p_headers, p_rows, [60, 40])

        # Category breakdown
        c_headers = ['Category', 'Total Amount']
        c_rows = [[k, f'${v:.2f}'] for k, v in sorted(category_totals.items())]
        add_table_section('Expense by Category', c_headers, c_rows, [60, 40])

        output = io.BytesIO()
        pdf.output(output)
        output.seek(0)
        return output, None
